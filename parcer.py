
import openpyxl
import xlsxwriter

class parcer:
    def __init__(self, name):

        self.name = "res\\" + str(name) + '.XLSX'

        # solution for issue - Export dataframe to xlsx - Error "zipfile.BadZipFile: File is not a zip file"
        # Create new XLSX file, using name from data search
        workbook = xlsxwriter.Workbook(self.name)
        workbook.close()
        
        self.wb = openpyxl.load_workbook(self.name, read_only = False)
        self.ws = self.wb.active
        #print("Fill header.")
        self.fill_data("№", 1, 1)
        self.fill_data("name", 1, 2)
        self.fill_data("address", 1, 3)
        self.fill_data("phone", 1, 4)
        self.fill_data("site", 1, 5)
        self.wb.save(self.name)
        #f.close()

    def parce(self, data, num):
        #print("Parcer def start")
        num_m = num + 1
        dd = data.split('\n')
        name = dd[0]
        address = ''
        phone = ''
        site = ''
        
        for i in dd:
            if i.endswith(".com") or i.endswith(".ca") or i.endswith(".biz") or i.endswith(".org") or i.endswith(".us"):
                site = i
            if ", FL " in i: # make it variable
                address = i
            if i.startswith("(") or i.replace(')', '').replace('(', '').replace('-', '').isdigit():
                phone = i

        #print(name, address, phone, site)
        self.fill_data(str(num_m), num, 1)
        self.fill_data(name, num_m, 2)
        self.fill_data(address, num_m, 3)
        self.fill_data(phone, num_m, 4)
        self.fill_data(site, num_m, 5)



    def fill_data(self, data, num, let):
        #print("Fill data and save")
        self.ws.cell(row=num, column=let, value = data)
        self.wb.save(self.name)


class parcer_simple:
    def __init__(self, name):

        self.name = "res\\" + str(name) + '.XLSX'

        # solution for issue - Export dataframe to xlsx - Error "zipfile.BadZipFile: File is not a zip file"
        # Create new XLSX file, using name from data search
        workbook = xlsxwriter.Workbook(self.name)
        workbook.close()
        
        self.wb = openpyxl.load_workbook(self.name, read_only = False)
        self.ws = self.wb.active
        #print("Fill header.")
        self.fill_data("№", 1, 1)
        self.fill_data("name", 1, 2)
        self.fill_data("address", 1, 3)
        self.fill_data("phone", 1, 4)
        self.fill_data("site", 1, 5)
        self.wb.save(self.name)
        #f.close()

    def parce(self, data):
        #print("Parcer def start")
        num_m = 1
        dd = data.split('\n')
        name = dd[0]
        address = ''
        phone = ''
        site = ''
        
        data_p = data.replace('Проложить маршрут', '\n').split('\n')
        
        #print("<<<      data found    >>>\n\n\n", data_p)
        counter = 1
        for i in data_p:
            self.fill_data(str(i), counter, 2)
            counter += 1
        

    def fill_data(self, data, num, let):
        #print("Fill data and save")
        self.ws.cell(row=num, column=let, value = data)
        self.wb.save(self.name)

